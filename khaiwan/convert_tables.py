from docx import Document
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def clean_cell(text):
    return text.strip().replace('\xa0', ' ') if text else ''

def is_header_row(row_cells):
    first = clean_cell(row_cells[0].text).lower()
    return 'ชื่อไฟล์' in first or 'file' in first

def extract_tables(docx_path, label):
    doc = Document(docx_path)
    results = []
    for i, table in enumerate(doc.tables):
        rows = []
        for row in table.rows:
            cells = [clean_cell(c.text) for c in row.cells]
            if rows and is_header_row(row.cells):
                continue
            rows.append(cells)
        results.append((f"{label}_Table{i+1}", rows))
    return results

pre_path  = r'C:\code\git\Senior_project\khaiwan\ตารางเปรียบเทียบการแปลงเสียงเป็นข้อมูลpre.docx'
post_path = r'C:\code\git\Senior_project\khaiwan\ตารางเปรียบเทียบการแปลงเสียงเป็นข้อมูลpost.docx'

all_sheets = extract_tables(pre_path, 'pre') + extract_tables(post_path, 'post')

wb = openpyxl.Workbook()
wb.remove(wb.active)

header_fill  = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
data_font    = Font(name='Calibri', size=11)
alt_fill     = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
thin   = Side(style='thin', color='BFBFBF')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for sheet_name, rows in all_sheets:
    ws = wb.create_sheet(title=sheet_name)
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            if r_idx == 1:
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = center_align
            else:
                cell.font      = data_font
                cell.fill      = alt_fill if r_idx % 2 == 0 else PatternFill()
                cell.alignment = center_align if c_idx > 1 else left_align
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)
    ws.freeze_panes = 'A2'

out_path = r'C:\code\git\Senior_project\khaiwan\ตารางเปรียบเทียบการแปลงเสียง.xlsx'
wb.save(out_path)
print('Saved:', out_path)
print('Sheets:', wb.sheetnames)
